import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import pandas as pd
import openpyxl
import requests
from PIL import Image
from io import BytesIO
from openpyxl.drawing.image import Image as ExcelImage

# WebDriver 초기화
options = webdriver.ChromeOptions()
options.add_argument("--headless")
options.add_argument("--window-size=1920,1080")

# driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)
driver.implicitly_wait(3)

# 함수
def get_movie_info(data):
  boxoffice_list_content = driver.find_element(By.ID, "boxoffice_list_content")
  boxoffice_li = boxoffice_list_content.find_elements(By.CLASS_NAME, "boxoffice_li")
  for li in boxoffice_li:
    rank = li.find_element(By.CLASS_NAME, "grade").text

    img_element = li.find_element(By.TAG_NAME, "img")
    img = img_element.get_attribute("src")

    mov_name = li.find_element(By.CLASS_NAME, "mov_name").text

    mov_name = li.find_element(By.CLASS_NAME, "mov_name").text

    pp_num = li.find_element(By.CLASS_NAME, "people_num").text

    link_element = li.find_element(By.TAG_NAME, "a")
    link = link_element.get_attribute("href")

    # print(f"{rank:<5}{mov_name} <{pp_num}>")
    data.append({"rank":rank, "image": img, "title":mov_name, "views":pp_num,  "link":link})

def get_current_page_idx():
  on_page_idx = int(driver.find_element(By.CSS_SELECTOR, '#boxoffice_list_content > div > div > a.on').text)
  print("-"*25, f"현재 페이지: {on_page_idx}", "-"*25)
  return on_page_idx
  
def get_next_page(on_page_idx):
  page_div = driver.find_element(By.CLASS_NAME, "page")
  page_btn = page_div.find_element(By.LINK_TEXT , str(on_page_idx+1))
  page_btn.click()
  time.sleep(2)

def img_convert(row_num, item):
  res = requests.get(item["image"])
  img = Image.open(BytesIO(res.content))
  img.thumbnail((100, 100)) # resize
  img_path = f"mov_imgs/image{row_num}.jpg"
  img.save(img_path)
  img_obj = ExcelImage(img_path)
  return img_obj

def to_excel_by_openpyxl(data):
  wb = openpyxl.Workbook()
  ws = wb.active
  ws.append(['Rank', "Poster", "Title", "Views", "Link"])

  for row_num, item in enumerate(data, start=2):
    img_obj = img_convert(row_num, item)
    ws.cell(row=row_num, column=1, value = item["rank"])
    ws.add_image(img_obj, f"B{row_num}")
    ws.cell(row=row_num, column=3, value = item["title"])
    ws.cell(row=row_num, column=4, value = item["views"])
    ws.cell(row=row_num, column=5, value = item["link"])
    ws.row_dimensions[row_num].height = 100     # 셀 높이 키우기
  wb.save("boxoffice_ranking_info.xlsx")

# 웹 페이지 열기
url = "http://www.cine21.com/rank/boxoffice/domestic"
driver.get(url)
page_source = driver.page_source

# # 첫 페이지만 엑셀로 넣기 
# data = []
# get_movie_info(data)
# to_excel_by_openpyxl(data)

# 앞에서 부터 순서대로 n개 페이지의 영화 가져오기 
required_page_number = 5
i = 0
data = []
while i < required_page_number:
  on_page_idx = get_current_page_idx()
  get_movie_info(data)
  if i != required_page_number-1:
    get_next_page(on_page_idx)
  else: 
    print("----------페이지 탐색이 종료 됩니다.----------")
  i += 1
to_excel_by_openpyxl(data)

# # to excel by pandas
# df = pd.DataFrame(data)
# df.to_excel("mov_data.xlsx", index = False)

# WebDriver 종료
driver.quit()